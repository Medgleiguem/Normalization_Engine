"""
Excel generator - creates normalized Excel files
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import List
from app.models.table_model import Table
from app.models.analysis_result import AnalysisResult

class ExcelGenerator:
    """Generate normalized Excel files from analysis results"""
    
    def __init__(self, analysis_result: AnalysisResult):
        self.analysis_result = analysis_result
        self.workbook = Workbook()
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
    
    def generate_excel(self, output_path: str) -> str:
        """Generate normalized Excel file"""
        # Add metadata sheet
        self._add_metadata_sheet()
        
        # Add each normalized table as a sheet
        for table in self.analysis_result.final_tables:
            self._add_table_sheet(table)
        
        # Add data dictionary
        self._add_data_dictionary()
        
        # Add relationships diagram (text-based)
        self._add_relationships_sheet()
        
        # Save workbook
        self.workbook.save(output_path)
        return output_path
    
    def _add_metadata_sheet(self):
        """Add metadata sheet with normalization information"""
        ws = self.workbook.create_sheet("Metadata", 0)
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        
        # Title
        ws['A1'] = 'Database Normalization Metadata'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Metadata
        metadata = [
            ('Original Table', self.analysis_result.original_table.name),
            ('Original Normal Form', self.analysis_result.current_normal_form.value),
            ('Target Normal Form', self.analysis_result.target_normal_form.value),
            ('Final Normal Form Achieved', self.analysis_result.normalization_steps[-1].to_nf.value if self.analysis_result.normalization_steps else 'N/A'),
            ('Number of Tables Created', len(self.analysis_result.final_tables)),
            ('Normalization Steps', len(self.analysis_result.normalization_steps)),
            ('Total Violations Resolved', len(self.analysis_result.get_all_violations())),
        ]
        
        row = 3
        for key, value in metadata:
            ws[f'A{row}'] = key
            ws[f'B{row}'] = str(value)
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
    
    def _add_table_sheet(self, table: Table):
        """Add a sheet for each normalized table"""
        ws = self.workbook.create_sheet(table.name)
        
        # Header styling
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # Add column headers
        for col_idx, column in enumerate(table.columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = column.name
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            
            # Mark primary key columns
            if column.name in table.primary_key:
                cell.value = f"{column.name} (PK)"
            
            # Mark foreign key columns
            if column.name in table.foreign_keys:
                ref_table, ref_col = table.foreign_keys[column.name]
                cell.value = f"{column.name} (FK → {ref_table})"
        
        # Add sample data if available
        if table.data:
            for row_idx, row_data in enumerate(table.data, 2):
                for col_idx, column in enumerate(table.columns, 1):
                    value = row_data.get(column.name, '')
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _add_data_dictionary(self):
        """Add data dictionary sheet"""
        ws = self.workbook.create_sheet("Data Dictionary")
        
        # Header
        headers = ['Table Name', 'Column Name', 'Data Type', 'Nullable', 'Unique', 'Key Type', 'Description']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        row = 2
        for table in self.analysis_result.final_tables:
            for column in table.columns:
                ws.cell(row=row, column=1, value=table.name)
                ws.cell(row=row, column=2, value=column.name)
                ws.cell(row=row, column=3, value=column.data_type.value)
                ws.cell(row=row, column=4, value='Yes' if column.nullable else 'No')
                ws.cell(row=row, column=5, value='Yes' if column.unique else 'No')
                
                # Key type
                key_type = ''
                if column.name in table.primary_key:
                    key_type = 'Primary Key'
                elif column.name in table.foreign_keys:
                    ref_table, ref_col = table.foreign_keys[column.name]
                    key_type = f'Foreign Key → {ref_table}({ref_col})'
                ws.cell(row=row, column=6, value=key_type)
                
                # Description
                description = f'Column from {table.name} table'
                ws.cell(row=row, column=7, value=description)
                
                row += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, 8):
            ws.column_dimensions[chr(64 + col_idx)].width = 20
    
    def _add_relationships_sheet(self):
        """Add relationships diagram sheet"""
        ws = self.workbook.create_sheet("Relationships")
        
        ws['A1'] = 'Table Relationships'
        ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        ws[f'A{row}'] = 'Table'
        ws[f'B{row}'] = 'Foreign Key'
        ws[f'C{row}'] = 'References'
        
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
            ws[f'{col}{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            ws[f'{col}{row}'].font = Font(color="FFFFFF", bold=True)
        
        row += 1
        
        for table in self.analysis_result.final_tables:
            if table.foreign_keys:
                for fk_col, (ref_table, ref_col) in table.foreign_keys.items():
                    ws[f'A{row}'] = table.name
                    ws[f'B{row}'] = fk_col
                    ws[f'C{row}'] = f'{ref_table}({ref_col})'
                    row += 1
        
        if row == 4:
            ws[f'A{row}'] = 'No foreign key relationships defined'
        
        # Column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
